

import re
import sys
import ast
import openpyxl
import warnings
import importlib
import pandas as pd
import importlib.util

warnings.filterwarnings("ignore")



class LOAD_TEST_DATA():
    def __init__(self, EXCEL_TEST_PATH):
        self.EXCEL_TEST_PATH                = EXCEL_TEST_PATH
        self.EXCEL                          = self.LOAD_TEST_DATA()

    def LOAD_TEST_DATA(self):
        TEST_WB                             = openpyxl.load_workbook(self.EXCEL_TEST_PATH)
        TEST_DICT_WB                        = {'SHEETS': {}, 'DATAFRAME': {}, 'COLUMN_MAP': {}}

        for idx, TEST_SHEET_NAME in enumerate(TEST_WB.sheetnames):
            TEST_WS                         = TEST_WB[TEST_SHEET_NAME]
            TEST_DF                         = pd.DataFrame(list(TEST_WS.values)[1:], columns=list(TEST_WS.values)[0])


            COLUMN_MAPPING_DF               = pd.DataFrame({  'LETTER' : [openpyxl.utils.get_column_letter(i + 1) for i in range(TEST_WS.max_column)],
                                                              'HEADER' : list(TEST_WS.values)[0]})

            TEST_DICT_WB['SHEETS'][idx]     = TEST_SHEET_NAME
            TEST_DICT_WB['DATAFRAME'][idx]  = TEST_DF
            TEST_DICT_WB['COLUMN_MAP'][idx] = COLUMN_MAPPING_DF

        return TEST_DICT_WB





def EXTRACTION(FORMULA):
    FORMULA = FORMULA.strip()

    PATTERN                 = r"([A-Z]+)\("
    FUNCTION_LIST           = re.findall(PATTERN, FORMULA)

    if not FUNCTION_LIST:   return ['HARDCODED']
    return FUNCTION_LIST





def NONE_PREFIX(VALUE):

    REFERENCE_PATTERN   = r"^[A-Za-z]+\d+$"
    RANGE_PATTERN       = r"^[A-Za-z]+:[A-Za-z]+$"
    
    if re.match(REFERENCE_PATTERN, VALUE) or re.match(RANGE_PATTERN, VALUE):

        if "!" not in VALUE:
            return f"NONE!{VALUE}"
        
    return VALUE



def PROCESS_CONDITIONS(CONDITIONS):

    CONDITION_PARTS     = re.split(r'(==|!=|<=|>=|<|>|=)', CONDITIONS)
    PROCESSED_PARTS     = [NONE_PREFIX(part.strip()) for part in CONDITION_PARTS]
    
    return ''.join(PROCESSED_PARTS)







def PARSE_FUNCTIONS(FORMULA):
    PATTERN             = r"([A-Z]+)\((.*)\)"
    MATCH               = re.match(PATTERN, FORMULA.strip())
    
    if MATCH:
        FUNCTION_NAME   = MATCH.group(1)
        ARGUMENT_STR    = MATCH.group(2).strip()
        
        ARGUMENTS       = SPLIT_ARGUMENTS(ARGUMENT_STR)
        
        return FUNCTION_NAME, ARGUMENTS
    
    return None, []






def SPLIT_ARGUMENTS(ARGUMENT_STR):
    ARGS, CURRENT_ARG, PARENTHESIS_COUNT    = [], "", 0

    for CHAR in ARGUMENT_STR:
        if CHAR == ',' and PARENTHESIS_COUNT == 0:
            ARGS.append(CURRENT_ARG.strip())
            CURRENT_ARG = ""

        else:
            CURRENT_ARG += CHAR
            if CHAR == '(':         PARENTHESIS_COUNT += 1
            elif CHAR == ')':       PARENTHESIS_COUNT -= 1

    if CURRENT_ARG:                 ARGS.append(CURRENT_ARG.strip())
    
    return ARGS




def PROCESS_FORMULA(FORMULA, FUNCTIONS_DF, FORMULA_LIST):

    TOP_FUNC_NAME, TOP_FUNC_ARGS    = PARSE_FUNCTIONS(FORMULA)
    if 'HARDCODED' in FORMULA_LIST: return JSON_OUTPUT_HARDCODED(FORMULA.strip('=').strip())
    if not TOP_FUNC_NAME:           return {}

    return JSON_OUTPUT(TOP_FUNC_NAME, TOP_FUNC_ARGS, FUNCTIONS_DF)



def JSON_OUTPUT_HARDCODED(value):
    return {"function"      : "HARDCODED",
            "components"    : {"STRING" : value}
            }






def JSON_OUTPUT(FUNCTION_NAME, ARGUMENTS, FUNCTIONS_DF):
    JSON_FIELDS                                             = (FUNCTIONS_DF[FUNCTIONS_DF['EXCEL_FUNCTION']==FUNCTION_NAME].reset_index().iloc[:,1:]).at[0, 'JSON_FIELDS']

    if len(ARGUMENTS) == len(JSON_FIELDS):
        COMPONENTS = {}
        
        for i in range(len(ARGUMENTS)):
            NESTED_FUNCS_MATCH                              = re.match(r"([A-Z]+)\((.*)\)", ARGUMENTS[i].strip())
            
            if NESTED_FUNCS_MATCH:
                NESTED_FUNCS_NAME                           = NESTED_FUNCS_MATCH.group(1)
                NESTED_FUNCS_ARGS                           = SPLIT_ARGUMENTS(NESTED_FUNCS_MATCH.group(2))
                COMPONENTS[JSON_FIELDS[i]]                  = JSON_OUTPUT(NESTED_FUNCS_NAME, NESTED_FUNCS_ARGS, FUNCTIONS_DF)

            else:
                if JSON_FIELDS[i] == "CONDITION":           COMPONENTS[JSON_FIELDS[i]] = PROCESS_CONDITIONS(ARGUMENTS[i].strip())
                else:                                       COMPONENTS[JSON_FIELDS[i]] = NONE_PREFIX(ARGUMENTS[i].strip())
        
        return {"function": FUNCTION_NAME, "components": COMPONENTS}
    
    else:    
        return {"function": FUNCTION_NAME, "components": ARGUMENTS}








def CALL_PYTHON_FUNCS(FUNCTION_NAME, FEATURES_DICT, WORKBOOK_DICT, FUNCTIONS_DF, FOLDER_PATH):

    if FUNCTION_NAME in FUNCTIONS_DF['PYTHON_FUNCTION'].values:
        PYTHON_FILE         = (FUNCTIONS_DF[FUNCTIONS_DF['EXCEL_FUNCTION'] == FUNCTION_NAME].reset_index().iloc[:, 1:]).at[0, 'PYTHON_FILE']
        PYTHON_FUNCTION     = (FUNCTIONS_DF[FUNCTIONS_DF['EXCEL_FUNCTION'] == FUNCTION_NAME].reset_index().iloc[:, 1:]).at[0, 'PYTHON_FUNCTION']

        sys.path.append(FOLDER_PATH)
        MODULE              = importlib.import_module(PYTHON_FILE)
        FUNC_TO_CALL        = getattr(MODULE, PYTHON_FUNCTION)

        return FUNC_TO_CALL(FEATURES_DICT, WORKBOOK_DICT)
    
    elif FUNCTION_NAME == "HARDCODED":return FEATURES_DICT['String'] 
    else:raise ValueError(f"Function {FUNCTION_NAME} not found in mapping.")
    



def GENERATE_PYTHON_SCRIPT(FORMULA_JSON, FUNCTIONS_DF, WORKBOOK_DICT, FOLDER_PATH):

    FUNCTION_NAME           = FORMULA_JSON["function"]
    COMPONENTS              = FORMULA_JSON["components"]
    FEATURES_DICT           = {} 

    for key, VALUE in COMPONENTS.items():
        if isinstance(VALUE, dict): FEATURES_DICT[key] = GENERATE_PYTHON_SCRIPT(VALUE, FUNCTIONS_DF, WORKBOOK_DICT, FOLDER_PATH)
        else: FEATURES_DICT[key] = VALUE

    return CALL_PYTHON_FUNCS(FUNCTION_NAME, FEATURES_DICT, WORKBOOK_DICT, FUNCTIONS_DF, FOLDER_PATH)




def SHEET_INDEXING(SHEET_NAME, DICT_INPUT):

    for index, NAME in DICT_INPUT['SHEETS'].items():
        if NAME == SHEET_NAME:
            return index
    raise ValueError(f"Sheet name '{SHEET_NAME}' not found in DICT_INPUT['SHEETS']")





def MAP_EXCEL_REF_PYTHON(REFERENCE, DICT_INPUT, OPERATORS_DF):


    if 'NONE!' in REFERENCE:
        REFERENCES                  = re.findall(r"NONE!\w+\d+", REFERENCE)

        for REF in REFERENCES:
            COL_LETTER              = re.search(r"NONE!(\w+)\d+", REF).group(1)
            DF_COL                  = DICT_INPUT['COLUMN_MAP'][0].set_index('LETTER').at[COL_LETTER, 'HEADER']
            REFERENCE               = REFERENCE.replace(REF, f"df['{DF_COL}']")


    elif re.search(r"\w+!\w+\d+", REFERENCE):
        REFERENCES                  = re.findall(r"\w+!\w+\d+", REFERENCE)

        for REF in REFERENCES:
            SHEET_NAME, CELL_REF    = REF.split('!')
            SHEET_INDEX             = SHEET_INDEXING(SHEET_NAME, DICT_INPUT)
            
            COL_LETTER              = re.search(r"(\w+)\d+", CELL_REF).group(1)
            COL_MAP_DF              = DICT_INPUT['COLUMN_MAP'][SHEET_INDEX]
            COL_NAME                = COL_MAP_DF.set_index('LETTER').at[COL_LETTER, 'HEADER']
            REFERENCE               = REFERENCE.replace(REF, f"WORKBOOK_DICT['DATAFRAME'][{SHEET_INDEX}]['{COL_NAME}']")


    for idx, ROW in OPERATORS_DF.iterrows():
        EXCEL_OP                    = ROW['EXCEL_OPERATOR']
        PYTHON_OP                   = ROW['PYTHON_OPERATOR']
        REFERENCE = re.sub(rf'(?<![a-zA-Z_0-9=]){EXCEL_OP}(?![a-zA-Z_0-9=])', PYTHON_OP, REFERENCE)

    return REFERENCE




def APPLY_OPERATORS(REFERENCE, OPERATORS_DF):

    INITIAL_EQUALS                  = REFERENCE.startswith('=')
    if INITIAL_EQUALS: REFERENCE    = REFERENCE[1:] 


    for idx, ROW in OPERATORS_DF.iterrows():
        EXCEL_OP                    = re.escape(ROW['EXCEL_OPERATOR'])
        PLACEHOLDER                 = ROW['PLACEHOLDER']
        REFERENCE                   = REFERENCE.replace(EXCEL_OP, PLACEHOLDER)
    

    for idx, ROW in OPERATORS_DF.iterrows():
        PLACEHOLDER                 = ROW['PLACEHOLDER']
        PYTHON_OP                   = ROW['PYTHON_OPERATOR']
        REFERENCE                   = REFERENCE.replace(PLACEHOLDER, PYTHON_OP)
    
    if INITIAL_EQUALS: REFERENCE    = '=' + REFERENCE

    return REFERENCE












class FORMULA_CONVERSION():
    def __init__(self, FORMULA, OPERATORS_DF, FUNCTIONS_DF, DICT_WORKBOOK, FOLDER_PATH):
        self.FORMULA                            = FORMULA
        self.OPERATORS_DF                       = OPERATORS_DF
        self.FUNCTIONS_DF                       = FUNCTIONS_DF
        self.DICT_WORKBOOK                      = DICT_WORKBOOK
        self.FOLDER_PATH                        = FOLDER_PATH
        self.JSON_FORMULA, self.PYTHON_SCRIPT   = self.FORMULA_CONVERSION()

    def FORMULA_CONVERSION(self):

        FORMULA_LIST                            = EXTRACTION(self.FORMULA)
        FORMULA_CLEAN                           = APPLY_OPERATORS(self.FORMULA, self.OPERATORS_DF)
        RESULT                                  = PROCESS_FORMULA(FORMULA_CLEAN[1:], self.FUNCTIONS_DF, FORMULA_LIST)
        PYTHON_SCRIPT                           = GENERATE_PYTHON_SCRIPT(RESULT, self.FUNCTIONS_DF, self.DICT_WORKBOOK, self.FOLDER_PATH)
        PYTHON_SCRIPT_CLEAN                     = MAP_EXCEL_REF_PYTHON(PYTHON_SCRIPT, self.DICT_WORKBOOK, self.OPERATORS_DF)


        return RESULT, PYTHON_SCRIPT_CLEAN






class FORMULA_JSON():
    def __init__(self, FORMULA, OPERATORS_DF, FUNCTIONS_DF):
        self.FORMULA                            = FORMULA
        self.OPERATORS_DF                       = OPERATORS_DF
        self.FUNCTIONS_DF                       = FUNCTIONS_DF
        self.JSON_FORMULA                       = self.FORMULA_JSON()

    def FORMULA_JSON(self):

        FORMULA_LIST                            = EXTRACTION(self.FORMULA)
        FORMULA_CLEAN                           = APPLY_OPERATORS(self.FORMULA, self.OPERATORS_DF)
        RESULT                                  = PROCESS_FORMULA(FORMULA_CLEAN[1:], self.FUNCTIONS_DF, FORMULA_LIST)

        return RESULT












class SOURCE_REFERENCE_INFORMATION():
    def __init__(self, EXCEL_FUNCS_PATH):
        self.EXCEL_FUNCS_PATH                   = EXCEL_FUNCS_PATH
        self.FUNCTIONS_DF, self.OPERATORS_DF    = self.SOURCE_REFERENCE_INFORMATION()

    def SOURCE_REFERENCE_INFORMATION(self):

        FUNC_DATA                               = LOAD_TEST_DATA(self.EXCEL_FUNCS_PATH)
        FUNCTIONS_DF                            = FUNC_DATA.EXCEL['DATAFRAME'][0]
        OPERATORS_DF                            = FUNC_DATA.EXCEL['DATAFRAME'][1]
        FUNCTIONS_DF['JSON_FIELDS']             = FUNCTIONS_DF['JSON_FIELDS'].apply(ast.literal_eval)

        return FUNCTIONS_DF, OPERATORS_DF




class SOURCE_TEST_INFORMATION():
    def __init__(self, EXCEL_TEST_PATH):
        self.EXCEL_TEST_PATH                    = EXCEL_TEST_PATH
        self.DICT_WORKBOOK, self.FORMULA        = self.SOURCE_TEST_INFORMATION()

    def SOURCE_TEST_INFORMATION(self):

        TEST_DATA                               = LOAD_TEST_DATA(self.EXCEL_TEST_PATH)
        FORMULA                                 = TEST_DATA.EXCEL['DATAFRAME'][0]['FORMULA'][0]

        return TEST_DATA.EXCEL, FORMULA








