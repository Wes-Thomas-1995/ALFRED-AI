import re
import ast
import zipfile
import openpyxl
import pandas as pd
import xml.etree.ElementTree as ET
from collections import defaultdict



class EXTRACT_FORMULAS:
    def __init__(self, WORKSHEETS):
        self.WORKSHEETS = WORKSHEETS
        self.DICT       = self.EXTRACT()

    def EXTRACT(self):

        FORMULAS = {}
        for SHEET_NAME, CELLS in self.WORKSHEETS.items():
            SHEET_FORMULAS = {}
            for CELL_REF, CELL_DATA in CELLS.items():
                if 'FORMULA' in CELL_DATA and CELL_DATA['FORMULA'] is not None:
                    FORMULA = CELL_DATA['FORMULA']
                    if any(err in FORMULA for err in ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A']):    SHEET_FORMULAS[CELL_REF] = {'FORMULA': FORMULA, 'ERROR': True}
                    else:                                                                                   SHEET_FORMULAS[CELL_REF] = {'FORMULA': FORMULA, 'ERROR': False}

            FORMULAS[SHEET_NAME] = SHEET_FORMULAS

        return FORMULAS
    

class CROSS_SHEET_REFERENCE_IDENTIFIER:
    def __init__(self, FORMULAS):
        self.FORMULAS = FORMULAS
        self.DICT = self.IDENTIFY()

    def IDENTIFY(self):
        CROSS_REFERENCES = {}

        # Regex pattern to identify sheet and cell references
        PATTERN = r'(\w+!)?([A-Z]+[0-9]+|[A-Z]+:[A-Z]+)'

        for SHEET_NAME, SHEET_FORMULAS in self.FORMULAS.items():
            # Initialize the sheet's reference list in CROSS_REFERENCES
            CROSS_REFERENCES[SHEET_NAME] = {}

            for CELL_REF, FORMULA_INFO in SHEET_FORMULAS.items():
                FORMULA = FORMULA_INFO['FORMULA']
                MATCHES = re.findall(PATTERN, FORMULA)
                
                if MATCHES:
                    REFERENCES = []
                    for match in MATCHES:
                        if match[0]:  # If the match contains a sheet name
                            REFERENCES.append({'SHEET': match[0].rstrip('!'), 'CELL': match[1]})
                        else:  # Intra-sheet reference; use the current SHEET_NAME
                            REFERENCES.append({'SHEET': SHEET_NAME, 'CELL': match[1]})
                    
                    # Add the references to the specific cell in CROSS_REFERENCES
                    CROSS_REFERENCES[SHEET_NAME][CELL_REF] = REFERENCES

        return CROSS_REFERENCES
    





class FORMULA_ANALYZER:
    def __init__(self, WORKSHEETS, FUNC_PATH):
        self.WORKSHEETS = WORKSHEETS
        self.FUNC_PATH  = FUNC_PATH


    def ANALYSE_FORMULAS(self):
        """
        Loop through each sheet and cell to find and parse formulas.
        Adds a breakdown of each formula to the cell data.
        """

        FUNCTIONS_DF, OPERATORS_DF = self.SOURCE_REFERENCE_INFORMATION(self.FUNC_PATH)

        for SHEET_NAME, CELLS in self.WORKSHEETS.items():
            for CELL_REFS, CELL_DATA in CELLS.items():
                FORMULA = CELL_DATA.get('FORMULA')
                if FORMULA:  
                    FORMULA_BREAKDOWN = self.FORMULA_CONVERSION(('='+FORMULA), OPERATORS_DF, FUNCTIONS_DF)
                    CELL_DATA['FORMULA_BREAKDOWN'] = FORMULA_BREAKDOWN



    def LOAD_TEST_DATA(self, EXCEL_FUNCS_PATH):
        TEST_WB                             = openpyxl.load_workbook(EXCEL_FUNCS_PATH)
        TEST_DICT_WB                        = {'SHEETS': {}, 'DATAFRAME': {}, 'COLUMN_MAP': {}}

        for idx, TEST_SHEET_NAME in enumerate(TEST_WB.sheetnames):
            TEST_WS                         = TEST_WB[TEST_SHEET_NAME]
            TEST_DF                         = pd.DataFrame(list(TEST_WS.values)[1:], columns=list(TEST_WS.values)[0])
            COLUMN_MAPPING_DF               = pd.DataFrame({'LETTER' : [openpyxl.utils.get_column_letter(i + 1) for i in range(TEST_WS.max_column)],
                                                            'HEADER' : list(TEST_WS.values)[0]})

            TEST_DICT_WB['SHEETS'][idx]     = TEST_SHEET_NAME
            TEST_DICT_WB['DATAFRAME'][idx]  = TEST_DF
            TEST_DICT_WB['COLUMN_MAP'][idx] = COLUMN_MAPPING_DF

        return TEST_DICT_WB



    def SOURCE_REFERENCE_INFORMATION(self, EXCEL_FUNCS_PATH):

        FUNC_DATA                       = self.LOAD_TEST_DATA(EXCEL_FUNCS_PATH)
        FUNCTIONS_DF                    = FUNC_DATA['DATAFRAME'][0]
        OPERATORS_DF                    = FUNC_DATA['DATAFRAME'][1]
        FUNCTIONS_DF['JSON_FIELDS']     = FUNCTIONS_DF['JSON_FIELDS'].apply(ast.literal_eval)
        FUNCTIONS_DF['CLUSTERS']        = FUNCTIONS_DF['CLUSTERS'].apply(ast.literal_eval)

        return FUNCTIONS_DF, OPERATORS_DF



    def FORMULA_CONVERSION(self, FORMULA, OPERATORS_DF, FUNCTIONS_DF):

        FORMULA_LIST        = self.EXTRACTION(FORMULA)
        FORMULA_CLEAN       = self.APPLY_OPERATORS(FORMULA, OPERATORS_DF)        
        RESULT              = self.PROCESS_FORMULA(FORMULA_CLEAN[1:], FUNCTIONS_DF, FORMULA_LIST)

        return RESULT



    def EXTRACTION(self, FORMULA):
        FORMULA = FORMULA.strip()

        PATTERN                 = r"([A-Z]+)\("
        FUNCTION_LIST           = re.findall(PATTERN, FORMULA)

        if not FUNCTION_LIST:   return ['HARDCODED']
        return FUNCTION_LIST




    def APPLY_OPERATORS(self, REFERENCE, OPERATORS_DF):

        INITIAL_EQUALS                  = REFERENCE.startswith('=')
        if INITIAL_EQUALS: REFERENCE    = REFERENCE[1:] 


        for idx, ROW in OPERATORS_DF.iterrows():
            EXCEL_OP                    = re.escape(ROW['EXCEL_OPERATOR'])
            PLACEHOLDER                 = ROW['PLACEHOLDER']
            REFERENCE                   = REFERENCE.replace(EXCEL_OP, PLACEHOLDER)
        

        for idx, ROW in OPERATORS_DF.iterrows():
            PLACEHOLDER                 = ROW['PLACEHOLDER']
            PYTHON_OP                   = ROW['PYTHON_OPERATOR']
            REFERENCE                   = REFERENCE.replace(PLACEHOLDER, PYTHON_OP)
        
        if INITIAL_EQUALS: REFERENCE    = '=' + REFERENCE

        return REFERENCE




    def PROCESS_FORMULA(self, FORMULA, FUNCTIONS_DF, FORMULA_LIST):

        TOP_FUNC_NAME, TOP_FUNC_ARGS    = self.PARSE_FUNCTIONS(FORMULA)
        if 'HARDCODED' in FORMULA_LIST: return self.JSON_OUTPUT_HARDCODED(FORMULA.strip('=').strip())
        if TOP_FUNC_NAME not in FUNCTIONS_DF['EXCEL_FUNCTION'].values:
            return {"FUNCTION"      : "UNKNOWN",
                    "COMPONENTS"    : { "FUNCTION_NAME" : TOP_FUNC_NAME,
                                        "ARGUMENTS"     : TOP_FUNC_ARGS}}
        
        if not TOP_FUNC_NAME:           return {}

        return self.JSON_OUTPUT(TOP_FUNC_NAME, TOP_FUNC_ARGS, FUNCTIONS_DF)


    def JSON_OUTPUT_HARDCODED(self, value):
        return {"FUNCTION"      : "HARDCODED",
                "COMPONENTS"    : {"STRING" : value}
                }



    def PARSE_FUNCTIONS(self, FORMULA):
        PATTERN             = r"([A-Z]+)\((.*)\)"
        MATCH               = re.match(PATTERN, FORMULA.strip())
        
        if MATCH:
            FUNCTION_NAME   = MATCH.group(1)
            ARGUMENT_STR    = MATCH.group(2).strip()
            
            ARGUMENTS       = self.SPLIT_ARGUMENTS(ARGUMENT_STR)
            
            return FUNCTION_NAME, ARGUMENTS
        
        return None, []




    def JSON_OUTPUT(self, FUNCTION_NAME, ARGUMENTS, FUNCTIONS_DF):
        JSON_FIELDS                                             = (FUNCTIONS_DF[FUNCTIONS_DF['EXCEL_FUNCTION']==FUNCTION_NAME].reset_index().iloc[:,1:]).at[0, 'JSON_FIELDS']

        if len(ARGUMENTS) == len(JSON_FIELDS):
            COMPONENTS = {}
            
            for i in range(len(ARGUMENTS)):
                NESTED_FUNCS_MATCH                              = re.match(r"([A-Z]+)\((.*)\)", ARGUMENTS[i].strip())
                
                if NESTED_FUNCS_MATCH:
                    NESTED_FUNCS_NAME                           = NESTED_FUNCS_MATCH.group(1)
                    NESTED_FUNCS_ARGS                           = self.SPLIT_ARGUMENTS(NESTED_FUNCS_MATCH.group(2))
                    COMPONENTS[JSON_FIELDS[i]]                  = self.JSON_OUTPUT(NESTED_FUNCS_NAME, NESTED_FUNCS_ARGS, FUNCTIONS_DF)

                else:
                    if JSON_FIELDS[i] == "CONDITION":           COMPONENTS[JSON_FIELDS[i]] = self.PROCESS_CONDITIONS(ARGUMENTS[i].strip())
                    else:                                       COMPONENTS[JSON_FIELDS[i]] = self.NONE_PREFIX(ARGUMENTS[i].strip())
            
            return {"FUNCTION": FUNCTION_NAME, "COMPONENTS": COMPONENTS}
        
        else:    
            return {"FUNCTION": FUNCTION_NAME, "COMPONENTS": ARGUMENTS}



    def SPLIT_ARGUMENTS(self, ARGUMENT_STR):
        ARGS, CURRENT_ARG, PARENTHESIS_COUNT    = [], "", 0

        for CHAR in ARGUMENT_STR:
            if CHAR == ',' and PARENTHESIS_COUNT == 0:
                ARGS.append(CURRENT_ARG.strip())
                CURRENT_ARG = ""

            else:
                CURRENT_ARG += CHAR
                if CHAR == '(':         PARENTHESIS_COUNT += 1
                elif CHAR == ')':       PARENTHESIS_COUNT -= 1

        if CURRENT_ARG:                 ARGS.append(CURRENT_ARG.strip())
        
        return ARGS




    def NONE_PREFIX(self, VALUE):

        REFERENCE_PATTERN   = r"^[A-Za-z]+\d+$"
        RANGE_PATTERN       = r"^[A-Za-z]+:[A-Za-z]+$"
        
        if re.match(REFERENCE_PATTERN, VALUE) or re.match(RANGE_PATTERN, VALUE):

            if "!" not in VALUE:
                return f"NONE!{VALUE}"
            
        return VALUE



    def PROCESS_CONDITIONS(self, CONDITIONS):

        CONDITION_PARTS     = re.split(r'(==|!=|<=|>=|<|>|=)', CONDITIONS)
        PROCESSED_PARTS     = [self.NONE_PREFIX(part.strip()) for part in CONDITION_PARTS]
        
        return ''.join(PROCESSED_PARTS)






class ADD_REFERENCES_TO_EXTRACTOR:

    def __init__(self, extractor_dict, cross_ref_dict):
        self.EXTRACTOR_DICT = extractor_dict
        self.CROSS_REF_DICT = cross_ref_dict

    def ADD_REFERENCES(self):
        for SHEET_NAME, CELLS in self.EXTRACTOR_DICT.items():
            for CELL_REF, CELL_DATA in CELLS.items():

                REFERENCES = []
                if SHEET_NAME in self.CROSS_REF_DICT and CELL_REF in self.CROSS_REF_DICT[SHEET_NAME]:
                    REFERENCES = self.CROSS_REF_DICT[SHEET_NAME][CELL_REF]
                
                CELL_DATA['REFERENCES'] = REFERENCES





class FORMULA_API:

    def __init__(self, WORKSHEETS, FUNC_PATH):
        self.WORKSHEETS = WORKSHEETS
        self.FUNC_PATH = FUNC_PATH
        self.DICT = self.FORMULA_API()

    
    def FORMULA_API(self):
        EXTRACTOR           = EXTRACT_FORMULAS(self.WORKSHEETS)
        FORMULA             = FORMULA_ANALYZER(EXTRACTOR.DICT, self.FUNC_PATH)
        FORMULA.ANALYSE_FORMULAS() 

        CROSS_REF           = CROSS_SHEET_REFERENCE_IDENTIFIER(EXTRACTOR.DICT)
        REFERENCES          = ADD_REFERENCES_TO_EXTRACTOR(EXTRACTOR.DICT, CROSS_REF.DICT)

        REFERENCES.ADD_REFERENCES()

        return EXTRACTOR.DICT











WORKSHEETS          = ''        # REPLACE WHEN IT COMES TO CALLING ALL FUNCTIONS.
FUNC_PATH           = r'/Users/westhomas/Desktop/ALFRED/1__REPORT_AUTOMATION/USEFUL_FUNCTIONS/FORMULA_API/REFERENCE_FILES/FUNCTIONS.xlsx'
FORMULA             = FORMULA_API(WORKSHEETS, FUNC_PATH)

FORMULA.DICT


