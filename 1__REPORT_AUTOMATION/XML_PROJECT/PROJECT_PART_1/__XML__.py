import re
import json
import ast
import zipfile
import string
import openpyxl
import pandas as pd
import networkx as nx
import xml.etree.ElementTree as ET
from collections import defaultdict
from openpyxl import load_workbook
from oletools.olevba import VBA_Parser







class EXCEL_XML_EXTRACTOR:
    def __init__(self, FILE_PATH):
        self.FILE_PATH = FILE_PATH
        self.XML_FILES = self.EXTRACT_XML_FILES()

    def EXTRACT_XML_FILES(self):
        XML_FILES = {}
        with zipfile.ZipFile(self.FILE_PATH, 'r') as ZIP_REF:
            for FILE in ZIP_REF.namelist():
                if FILE.endswith('.xml'):
                    with ZIP_REF.open(FILE) as f:
                        XML_FILES[FILE] = f.read()
        return XML_FILES








class SHEET_NAME_MAP:
    def __init__(self, FILE_PATH):
        self.FILE_PATH = FILE_PATH
        self.CODE = self.GENERATE_SHEET_NAME_MAP()

    def VISIBLE_SHEE_MAP(self):
        with zipfile.ZipFile(self.FILE_PATH, 'r') as ZIP_REF:
            with ZIP_REF.open('xl/workbook.xml') as WORKBOOK_XML:
                TREE = ET.parse(WORKBOOK_XML)
                TREE_ROOT = TREE.getroot()
                NS = {'n': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                SHEET_NAMES = [SHEET.attrib['name'] for SHEET in TREE_ROOT.findall('.//n:sheets/n:sheet', NS)]
                return SHEET_NAMES

    def CODE_NAMES(self):
        CODE_NAMES, CLEAN_NAME, MOD_NAMES, MOD_REF = [], [], [], []
        VBA_NAME = 'xl/vbaProject.bin'
        VBA_TEMP = 'vbaProject.bin'

        with zipfile.ZipFile(self.FILE_PATH, 'r') as ZIP_REF:
            if VBA_NAME in ZIP_REF.namelist():
                with open(VBA_TEMP, 'wb') as VBA_FILE:
                    VBA_FILE.write(ZIP_REF.read(VBA_NAME))
                VBA_PARSER = VBA_Parser(VBA_TEMP)
                if VBA_PARSER.detect_vba_macros():
                    for (_, _, VBA_FILENAME, _) in VBA_PARSER.extract_macros():
                        if "ThisWorkbook" not in VBA_FILENAME and VBA_FILENAME.endswith('.cls'):
                            CODE_NAMES.append(VBA_FILENAME)
                            CLEAN_CODE_NAME = VBA_FILENAME.split('.cls')[0]
                            CLEAN_NAME.append(CLEAN_CODE_NAME)
                        if "ThisWorkbook" not in VBA_FILENAME and VBA_FILENAME.endswith('.bas'):
                            MOD_NAMES.append(VBA_FILENAME)
                            MOD_REF_1 = VBA_FILENAME.split('.bas')[0]
                            MOD_REF.append(MOD_REF_1)
        return CODE_NAMES, CLEAN_NAME, MOD_NAMES, MOD_REF

    def GENERATE_SHEET_NAME_MAP(self):
        if self.FILE_PATH.endswith('.xlsm'):
            VIS_SHEET_NAME = self.VISIBLE_SHEE_MAP()
            CODE_NAMES, CLEAN_NAME, MOD_NAMES, MOD_REF = self.CODE_NAMES()

            # Adjust the lengths of lists to be the same by padding shorter lists with 'NA'
            max_length = max(len(VIS_SHEET_NAME), len(CLEAN_NAME))

            VIS_SHEET_NAME += ['NA'] * (max_length - len(VIS_SHEET_NAME))
            CLEAN_NAME += ['NA'] * (max_length - len(CLEAN_NAME))
            CODE_NAMES += ['NA'] * (max_length - len(CODE_NAMES))

            # Create DataFrames for class and standard modules
            SHEET_DF = pd.DataFrame({
                'MODULE_TYPE': 'Class Module',
                'SHEET_NAME': VIS_SHEET_NAME,
                'VBA_CLEAN_SHEET_NAME': CLEAN_NAME,
                'VBA_SHEET_NAME': CODE_NAMES
            })

            MOD_DF = pd.DataFrame({
                'MODULE_TYPE': 'Standard Module',
                'SHEET_NAME': ['NA'] * len(MOD_REF),
                'VBA_CLEAN_SHEET_NAME': MOD_REF,
                'VBA_SHEET_NAME': MOD_NAMES
            })

            # Combine both DataFrames and return the result
            return pd.concat([SHEET_DF, MOD_DF], ignore_index=True)

        else:
            return "NOT VBA FILE"



class VBA_CODE:
    def __init__(self, FILE_PATH):
        self.FILE_PATH      = FILE_PATH
        self.EXTRACT        = []
        self.CODE           = self.VBA_SOURCE()

    def SPLIT_VBA(self, VBA_CODE):
        VBA_STARTER = ["Sub", "Private Sub", "Function", "Private Function"]
        VBA_ENDER = ["End Sub", "End Function"]
        PATTERN = r'(?P<block>(' + '|'.join(re.escape(starter) for starter in VBA_STARTER) + r').*?(' + '|'.join(re.escape(ender) for ender in VBA_ENDER) + r'))'
        MATCHES = re.finditer(PATTERN, VBA_CODE, re.DOTALL | re.MULTILINE)
        return [MATCH.group('block').strip() for MATCH in MATCHES]

    def VBA_SOURCE(self):
        VBA_NAME = 'xl/vbaProject.bin'
        VBA_TEMP = 'vbaProject.bin'

        with zipfile.ZipFile(self.FILE_PATH, 'r') as EXCEL_ZIP:
            if VBA_NAME in EXCEL_ZIP.namelist():
                with open(VBA_TEMP, 'wb') as VBA_FILE:
                    VBA_FILE.write(EXCEL_ZIP.read(VBA_NAME))

                VBA_PARSER = VBA_Parser(VBA_TEMP)
                if VBA_PARSER.detect_vba_macros():
                    for (_, _, VBA_FILENAME, VBA_CODE) in VBA_PARSER.extract_macros():
                        if VBA_CODE.strip():
                            if any(STARTER in VBA_CODE for STARTER in ["Sub", "Private Sub", "Function", "Private Function"]):
                                SPLIT_LIST = self.SPLIT_VBA(VBA_CODE)
                                for ITEM in SPLIT_LIST:
                                    self.EXTRACT.append({   'VBA_FILENAME'  : VBA_FILENAME,
                                                            'VBA_CODE'      : ITEM})
                                    

        return pd.DataFrame(self.EXTRACT)







class WORKBOOK_PARSING:
    def __init__(self, XML_FILES, LABELS):
        self.XML_FILES = XML_FILES
        self.LABELS = LABELS
        self.DICT = self.PARSE_WORKBOOK()

    def PARSE_WORKBOOK(self):
        MAP = {}
        if 'xl/workbook.xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/workbook.xml'])
            SHEETS = ROOT.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
            for i, SHEET in enumerate(SHEETS):
                NAME = SHEET.attrib.get('name')
                STATE = SHEET.attrib.get('state', 'visible')
                SHEET_ID = SHEET.attrib.get('sheetId')
                MAP[f'sheet{i + 1}'] = {'NAME': NAME, 'STATE': STATE, 'ID': SHEET_ID}
        
        for sheet_key, sheet_data in MAP.items():
            sheet_name = sheet_data['NAME']  # Get the sheet name from sheet_name_map
            if sheet_name in self.LABELS:
                sheet_data['LABEL'] = self.LABELS[sheet_name]  # Add the corresponding label

        return MAP

    # Method to map a table XML file to the corresponding sheet
    def get_sheet_name_by_table(self, FILE_NAME):
        for sheet_key, sheet_data in self.DICT.items():
            if sheet_key in FILE_NAME:
                return sheet_data['NAME']
        return None









class WORKSHEET_PARSING:
    def __init__(self, XML_FILES, SHEET_NAME_MAP):
        self.XML_FILES = XML_FILES
        self.SHEET_NAME_MAP = SHEET_NAME_MAP
        self.DICT = self.PARSE_WORKSHEETS()

    def PARSE_WORKSHEETS(self):
        WORKSHEETS = {}
        for FILE_NAME, XML_CONTENT in self.XML_FILES.items():
            if 'xl/worksheets/sheet' in FILE_NAME:
                ROOT = ET.fromstring(XML_CONTENT)
                SHEET_DATA = {}

                for CELL in ROOT.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    CELL_REF = CELL.attrib.get('r')
                    CELL_TYPE = CELL.attrib.get('t')
                    STYLE_INDEX = CELL.attrib.get('s')
                    FORMULA = CELL.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f')
                    VALUE = CELL.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')

                    VALUE_TEXT = VALUE.text if VALUE is not None else None

                    CELL_DATA = {
                        'TYPE': CELL_TYPE,
                        'STYLE_INDEX': STYLE_INDEX,
                        'FORMULA': FORMULA.text if FORMULA is not None else None,
                        'VALUE': VALUE_TEXT
                    }

                    SHEET_DATA[CELL_REF] = CELL_DATA

                # Extract the sheet internal reference (e.g., 'sheet1')
                SHEET_NAME_KEY = FILE_NAME.split('/')[-1].replace('.xml', '')
                # Map the internal sheet name to the user-defined sheet name
                USER_DEFINED_NAME = self.SHEET_NAME_MAP.get(SHEET_NAME_KEY, {}).get('NAME', SHEET_NAME_KEY)
                WORKSHEETS[USER_DEFINED_NAME] = SHEET_DATA

        return WORKSHEETS





class CALC_CHAIN_PARSING:
    def __init__(self, XML_FILES, SHEET_NAME_MAP):
        self.XML_FILES = XML_FILES
        self.SHEET_NAME_MAP = SHEET_NAME_MAP
        self.DICT = self.CALC_CHAIN_PARSE()

    def CALC_CHAIN_PARSE(self):
        CALC_CHAIN = []
        if 'xl/calcChain.xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/calcChain.xml'])
            for CELL in ROOT.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                CELL_REF = CELL.attrib.get('r')
                SHEET_ID = CELL.attrib.get('i')
                # Map SHEET_ID to the user-defined sheet name using SHEET_NAME_MAP
                SHEET_NAME = self.SHEET_NAME_MAP.get(f'sheet{SHEET_ID}', {}).get('NAME', f'sheet{SHEET_ID}')
                CALC_CHAIN.append({'CELL_REF': CELL_REF, 'SHEET_NAME': SHEET_NAME})
        return CALC_CHAIN





class SHARED_STRING_PARSING:

    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.SHARED_STRINGS_PARSE()

    def SHARED_STRINGS_PARSE(self):
        SHARED_STRINGS = {}
        if 'xl/sharedStrings.xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/sharedStrings.xml'])
            for i, SI in enumerate(ROOT.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si')):
                TEXT_NODE = SI.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                if TEXT_NODE is not None:
                    SHARED_STRINGS[i] = TEXT_NODE.text
        return SHARED_STRINGS








class STYLES_PARSING:
    
    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.STYLE_PARSE()

    def STYLE_PARSE(self):
        STYLES = {  
            'NUMBER_FORMATS': {},
            'FONTS': [],
            'FILLS': [],
            'BORDERS': [],
            'CELLXFS': []
        }
        
        if 'xl/styles.xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/styles.xml'])

            # 1. Extract number formats
            NUM_FMTS = ROOT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}numFmts')
            if NUM_FMTS is not None:
                for NUM_FMT in NUM_FMTS.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}numFmt'):
                    FMT_ID = NUM_FMT.attrib.get('numFmtId')
                    FORMAT_CODE = NUM_FMT.attrib.get('formatCode')
                    STYLES['NUMBER_FORMATS'][FMT_ID] = FORMAT_CODE

            # 2. Extract fonts
            FONTS = ROOT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}fonts')
            if FONTS is not None:
                for FONT in FONTS.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}font'):
                    FONT_DATA = {
                        'NAME': FONT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}name').attrib.get('val'),
                        'SIZE': FONT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sz').attrib.get('val'),
                        'BOLD': FONT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}b') is not None,
                        'ITALIC': FONT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}i') is not None,
                        'COLOR': FONT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}color').attrib.get('rgb') if FONT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}color') is not None else None
                    }
                    STYLES['FONTS'].append(FONT_DATA)

            # 3. Extract fills
            FILLS = ROOT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}fills')
            if FILLS is not None:
                for FILL in FILLS.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}fill'):
                    PATTERN_FILL = FILL.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}patternFill')
                    FILL_DATA = {
                        'PATTERN_TYPE': PATTERN_FILL.attrib.get('patternType') if PATTERN_FILL is not None else None,
                        'FG_COLOR': PATTERN_FILL.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}fgColor').attrib.get('rgb') if PATTERN_FILL is not None and PATTERN_FILL.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}fgColor') is not None else None,
                        'BG_COLOR': PATTERN_FILL.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}bgColor').attrib.get('rgb') if PATTERN_FILL is not None and PATTERN_FILL.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}bgColor') is not None else None
                    }
                    STYLES['FILLS'].append(FILL_DATA)

            # 4. Extract borders
            BORDERS = ROOT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}borders')
            if BORDERS is not None:
                for BORDER in BORDERS.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}border'):
                    BORDER_DATA = {
                        'LEFT': BORDER.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}left') is not None,
                        'RIGHT': BORDER.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}right') is not None,
                        'TOP': BORDER.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}top') is not None,
                        'BOTTOM': BORDER.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}bottom') is not None
                    }
                    STYLES['BORDERS'].append(BORDER_DATA)

            # 5. Extract cellXfs (Cell formats)
            CELLXFS = ROOT.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cellXfs')
            if CELLXFS is not None:
                for XF in CELLXFS.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}xf'):
                    XF_DATA = {
                        'NUM_FMT_ID': XF.attrib.get('numFmtId'),
                        'FONT_ID': XF.attrib.get('fontId'),
                        'FILL_ID': XF.attrib.get('fillId'),
                        'BORDER_ID': XF.attrib.get('borderId')
                    }
                    STYLES['CELLXFS'].append(XF_DATA)

        return STYLES









class DRAWING_PARSING:

    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.DRAWING_PARSE()

    def DRAWING_PARSE(self):
        DRAWINGS = defaultdict(list)
        for FILE_NAME, XML_CONTENT in self.XML_FILES.items():
            if 'xl/drawings/drawing' in FILE_NAME:
                ROOT = ET.fromstring(XML_CONTENT)
                DRAWING_INFO = []


                # Look for twoCellAnchor elements
                for ANCHOR in ROOT.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor'):


                    # Identify images
                    BLIP = ANCHOR.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
                    if BLIP is not None:
                        EMBED = BLIP.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        DRAWING_INFO.append({'TYPE': 'IMAGE', 'EMBED_ID': EMBED})

                    # Identify shapes (e.g., rectangles, circles)
                    SP = ANCHOR.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}sp')
                    if SP is not None:
                        SP_NAME = SP.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}nvSpPr/{http://schemas.openxmlformats.org/drawingml/2006/main}cNvPr').attrib.get('name', 'Shape')
                        DRAWING_INFO.append({'TYPE': 'SHAPE', 'NAME': SP_NAME})

                    # Identify charts
                    GRAPHIC_FRAME = ANCHOR.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}graphicFrame')
                    if GRAPHIC_FRAME is not None:
                        CHART = GRAPHIC_FRAME.find('.//{http://schemas.openxmlformats.org/drawingml/2006/chart}chart')
                        if CHART is not None:
                            CHART_ID = CHART.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                            DRAWING_INFO.append({'TYPE': 'CHART', 'CHART_ID': CHART_ID})

                    # Identify connectors (e.g., lines)
                    CNX_SP = ANCHOR.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}cxnSp')
                    if CNX_SP is not None:
                        CNX_NAME = CNX_SP.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}nvCxnSpPr/{http://schemas.openxmlformats.org/drawingml/2006/main}cNvPr').attrib.get('name', 'Connector')
                        DRAWING_INFO.append({'TYPE': 'CONNECTOR', 'NAME': CNX_NAME})

                DRAWINGS[FILE_NAME] = DRAWING_INFO
        return DRAWINGS





class THEME_PARSING:

    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.THEME_PARSE()

    def THEME_PARSE(self):

        THEME = {}
        if 'xl/theme/theme1.xml' in self.XML_FILES:  
            ROOT = ET.fromstring(self.XML_FILES['xl/theme/theme1.xml'])
            COLOUR_SCHEME = []
            for COLOUR in ROOT.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/main}clrScheme//{http://schemas.openxmlformats.org/drawingml/2006/main}srgbClr'):
                COLOUR_SCHEME.append(COLOUR.attrib.get('val'))
            THEME['color_scheme'] = COLOUR_SCHEME
        return THEME




class RELATIVE_SHEETS_PARSING:
    
    def __init__(self, XML_FILES, SHEET_NAME_MAP):
        self.XML_FILES = XML_FILES
        self.SHEET_NAME_MAP = SHEET_NAME_MAP
        self.DICT = self.parse_sheet_rels()

    def parse_sheet_rels(self):
        SHEET_RELS = {}

        for FILE_NAME, XML_CONTENT in self.XML_FILES.items():
            if 'xl/worksheets/_rels/sheet' in FILE_NAME and '.rels' in FILE_NAME:
                try:    ROOT = ET.fromstring(XML_CONTENT)
                except ET.ParseError as e:
                    print(f"XML Parsing Error in {FILE_NAME}: {e}")
                    continue


                SHEET_NUM_KEY = FILE_NAME.split('/')[-1].replace('.xml.rels', '')
                USER_DEFINED_NAME = self.SHEET_NAME_MAP.get(SHEET_NUM_KEY, {}).get('NAME', SHEET_NUM_KEY)
                SHEET_RELS[USER_DEFINED_NAME] = []

                for REL in ROOT.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    REL_ID = REL.attrib.get('Id')
                    TARGET = REL.attrib.get('Target')
                    REL_TYPE = REL.attrib.get('Type').split('/')[-1].upper() 
                    
                    SHEET_RELS[USER_DEFINED_NAME].append({
                        'ID': REL_ID,
                        'TARGET': TARGET,
                        'TYPE': REL_TYPE
                    })
                    
        return SHEET_RELS



class CONTENT_TYPE_PARSING:

    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.CONTENT_TYPE_PARSE()

    def CONTENT_TYPE_PARSE(self):

        CONTENT_TYPES = []
        if '[Content_Types].xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['[Content_Types].xml'])
            for OVERRIDE in ROOT.findall('.//{http://schemas.openxmlformats.org/package/2006/content-types}Override'):
                PART_NAME       = OVERRIDE.attrib.get('PartName')
                CONTENT_TYPE    = OVERRIDE.attrib.get('ContentType')
                CONTENT_TYPES.append({'PART_NAME': PART_NAME, 'CONTENT_TYPE': CONTENT_TYPE})
        return CONTENT_TYPES





class RELATIVE_WORKBOOK_PARSING:
    
    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.WORKBOOK_REL_PARSE()

    def WORKBOOK_REL_PARSE(self):

        WORKBOOK_RELS = []
        if 'xl/_rels/workbook.xml.rels' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/_rels/workbook.xml.rels'])
            for REL in ROOT.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                REL_ID      = REL.attrib.get('Id')
                TARGET      = REL.attrib.get('Target')
                REL_TYPE    = REL.attrib.get('Type').split('/')[-1]
                WORKBOOK_RELS.append({'ID': REL_ID, 'TARGET': TARGET, 'TYPE': REL_TYPE})
        return WORKBOOK_RELS





class PIVOT_DATA:
    def __init__(self, FILE_PATH):
        self.FILE_PATH = FILE_PATH
        self.DICT = self.PIVOT_INFO()

    def PIVOT_INFO(self):
        WB                  = load_workbook(self.FILE_PATH, data_only=True)
        PIVOT_DATA          = {}

        for SHEET_NAME in WB.sheetnames:
            WS              = WB[SHEET_NAME]
            PIVOT_TABLES    = getattr(WS, '_pivots', None)

            if PIVOT_TABLES:
                for PIVOT in PIVOT_TABLES:
                    CACHE_DF                    = self.GET_CACHE(PIVOT)
                    FIELDS_DICT                 = self.GET_FIELDS(PIVOT, PIVOT.cache)
                    PIVOT_INFO                  = {}
                    PIVOT_INFO['NAME']          = PIVOT.name
                    PIVOT_INFO['SHEET']         = SHEET_NAME
                    PIVOT_INFO['LOCATION']      = PIVOT.location.ref
                    PIVOT_INFO['TABLE']         = { 'ROWS'      :       FIELDS_DICT['ROWS'],
                                                    'COLUMNS'   :       FIELDS_DICT['COLUMNS'],
                                                    'DATA'      :       FIELDS_DICT['DATA'],
                                                    'FILTERS'   :       FIELDS_DICT['FILTERS']}
                    PIVOT_INFO['SOURCE_TABLE']  = (PIVOT.cache.cacheSource.worksheetSource.name is not None)
                    PIVOT_INFO['SOURCE_INFO'] = {'NAME'         : PIVOT.cache.cacheSource.worksheetSource.name,
                                                'SHEET'        : PIVOT.cache.cacheSource.worksheetSource.sheet,
                                                'REF'          : PIVOT.cache.cacheSource.worksheetSource.ref,
                                                'DB'           : CACHE_DF}
                    PIVOT_INFO['DATA_FIELD'] = {'SUBTOTAL_TYPE' : [PIVOT.dataFields[A].subtotal for A in range(len(PIVOT.dataFields))] if isinstance(PIVOT.dataFields, list) else [PIVOT.dataFields.subtotal],
                                                'NAME'          : [PIVOT.dataFields[A].name for A in range(len(PIVOT.dataFields))] if isinstance(PIVOT.dataFields, list) else [PIVOT.dataFields.name],
                                                'ID'            : [PIVOT.dataFields[A].fld for A in range(len(PIVOT.dataFields))] if isinstance(PIVOT.dataFields, list) else [PIVOT.dataFields.fld]}

                    PIVOT_DATA[PIVOT.name]      = PIVOT_INFO
        
        return PIVOT_DATA

    def GET_CACHE(self, PIVOT_TABLE):
        FIELDS_MAP = {}
        for FIELD in PIVOT_TABLE.cache.cacheFields:
            if FIELD.sharedItems.count > 0:
                L = []
                for F in FIELD.sharedItems._fields:
                    try:
                        L += [F.v]
                    except AttributeError:
                        L += [""]
                FIELDS_MAP[FIELD.name] = L

        COLUMN_NAMES    = [FIELD.name for FIELD in PIVOT_TABLE.cache.cacheFields]
        ROWS            = []

        for RECORD in PIVOT_TABLE.cache.records.r:
            RECORD_VALUES   = [FIELD.v for FIELD in RECORD._fields]
            ROW_DICT        = {K: V for K, V in zip(COLUMN_NAMES, RECORD_VALUES)}

            for KEY in FIELDS_MAP:
                ROW_DICT[KEY] = FIELDS_MAP[KEY][ROW_DICT[KEY]]
            ROWS.append(ROW_DICT)

        DF = pd.DataFrame.from_dict(ROWS)
        return DF

    def GET_FIELDS(self, PIVOT_INFO, CACHE_INFO):
        CACHE_FIELDS = CACHE_INFO.cacheFields  # Cache fields

        # Extract row fields (from rowFields, using the 'x' attribute)
        if hasattr(PIVOT_INFO, 'rowFields'):
            ROW_FIELD_INDICES   = [FIELD.x for FIELD in PIVOT_INFO.rowFields]
            ROW_FIELD_NAMES     = [CACHE_FIELDS[IDX].name for IDX in ROW_FIELD_INDICES]
        else:
            ROW_FIELD_NAMES    = []

        # Extract column fields (from colFields, using the 'x' attribute)
        if hasattr(PIVOT_INFO, 'colFields'):
            COL_FIELD_INDICES   = [FIELD.x for FIELD in PIVOT_INFO.colFields]
            COL_FIELD_NAMES     = [CACHE_FIELDS[IDX].name for IDX in COL_FIELD_INDICES]
        else:
            COL_FIELD_NAMES    = []

        # Extract data fields (from dataFields, using the 'fld' attribute)
        if hasattr(PIVOT_INFO, 'dataFields'):
            DATA_FIELD_INDICES  = [FIELD.fld for FIELD in PIVOT_INFO.dataFields]
            DATA_FIELD_NAMES    = [CACHE_FIELDS[IDX].name for IDX in DATA_FIELD_INDICES]
        else:
            DATA_FIELD_NAMES   = []

        # Check for filters (from pageFields, using the 'fld' attribute)
        FILTER_INFO = []
        if hasattr(PIVOT_INFO, 'pageFields'):
            for PAGE_FIELD in PIVOT_INFO.pageFields:
                FILTER_INDEX    = PAGE_FIELD.fld
                FILTER_NAME     = CACHE_FIELDS[FILTER_INDEX].name
                FILTER_ITEM     = PAGE_FIELD.item  # The selected filter value

                if FILTER_ITEM is not None:
                    FILTER_VALUE = CACHE_FIELDS[FILTER_INDEX].sharedItems._fields[FILTER_ITEM].v  # Value from cache
                else:
                    FILTER_VALUE = "All"  # No specific filter applied, 'All' selected

                FILTER_INFO.append({'FILTER_NAME'   : FILTER_NAME,
                                    'FILTER_VALUE'  : FILTER_VALUE})

        # Return the field names for rows, columns, data fields, and any filters
        return {'ROWS'          : ROW_FIELD_NAMES,
                'COLUMNS'       : COL_FIELD_NAMES,
                'DATA'          : DATA_FIELD_NAMES,
                'FILTERS'       : FILTER_INFO}








class EXTRACT_DATA_MODELS:
    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.EXTRACT_MODEL()
    
    def EXTRACT_MODEL(self):
        DATA_MODELS = {'TABLES': [], 'RELATIONSHIPS': []}
        
        if 'xl/model/tables.xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/model/tables.xml'])
            for TABLE_NODE in ROOT.findall('.//{http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac}table'):
                TABLE_INFO = {
                    'NAME': TABLE_NODE.attrib.get('name'),
                    'ID': TABLE_NODE.attrib.get('id'),
                    'LOCATION': 'xl/model/tables.xml',
                    'DIMENSIONS': TABLE_NODE.attrib.get('ref')
                }
                DATA_MODELS['TABLES'].append(TABLE_INFO)

        if 'xl/model/relationships.xml' in self.XML_FILES:
            ROOT = ET.fromstring(self.XML_FILES['xl/model/relationships.xml'])
            for REL_NODE in ROOT.findall('.//{http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac}relationship'):
                RELATIONSHIP_INFO = {
                    'ID': REL_NODE.attrib.get('id'),
                    'SOURCE_TABLE': REL_NODE.attrib.get('source'),
                    'TARGET_TABLE': REL_NODE.attrib.get('target'),
                    'TYPE': REL_NODE.attrib.get('type')
                }
                DATA_MODELS['RELATIONSHIPS'].append(RELATIONSHIP_INFO)

        return DATA_MODELS


class TABLES_PARSING:
    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.TABLES_PARSE()

    def TABLES_PARSE(self):
        TABLES = {}
        for FILE_NAME, XML_CONTENT in self.XML_FILES.items():
            if 'xl/tables/' in FILE_NAME:
                ROOT = ET.fromstring(XML_CONTENT)
                TABLE_INFO = {}

                TABLE_INFO['NAME'] = ROOT.attrib.get('name')
                TABLE_INFO['REF'] = ROOT.attrib.get('ref') 
                TABLE_INFO['LOCATION'] = "NOPE"

                if TABLE_INFO['REF']:
                    start_cell, end_cell = TABLE_INFO['REF'].split(':')
                    start_row = int(re.findall(r'\d+', start_cell)[0])
                    end_row = int(re.findall(r'\d+', end_cell)[0])
                    row_count = end_row - start_row + 1
                    col_count = ord(end_cell[0]) - ord(start_cell[0]) + 1
                    TABLE_INFO['SHEET_LOCATION'] = start_cell
                    TABLE_INFO['DIMENSIONS'] = {"ROWS" : row_count,
                                                "COLS" : col_count}


                TABLE_INFO['COLUMNS'] = [col.attrib.get('name') for col in ROOT.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}tableColumn')]

                TABLES[FILE_NAME] = TABLE_INFO

        return TABLES

class EXTERNAL_CONNECTIONS_PARSING:

    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.DICT = self.CHECK_EXTERNAL_CONNECTIONS()

    def CHECK_EXTERNAL_CONNECTIONS(self):
        external_connections = {}

        # Check for connections.xml file that contains external connections
        if 'xl/connections.xml' in self.XML_FILES:
            root = ET.fromstring(self.XML_FILES['xl/connections.xml'])
            
            for connection in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}connection'):
                conn_id = connection.attrib.get('id')
                conn_name = connection.attrib.get('name')
                conn_type = connection.attrib.get('type')
                conn_ref = connection.attrib.get('ref', 'N/A')

                external_connections[conn_id] = {
                    'NAME': conn_name,
                    'TYPE': conn_type,
                    'REF': conn_ref
                }

        # Check workbook relationships for external link
        if 'xl/_rels/workbook.xml.rels' in self.XML_FILES:
            root = ET.fromstring(self.XML_FILES['xl/_rels/workbook.xml.rels'])

            for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                rel_type = rel.attrib.get('Type')
                rel_target = rel.attrib.get('Target')

                # Check if the relationship is external
                if 'externalLink' in rel_type:
                    external_connections[rel.attrib.get('Id')] = {
                        'NAME': rel.attrib.get('Target'),
                        'TYPE': 'External Link',
                        'REF': rel.attrib.get('Target')
                    }

        return external_connections
    

class POWER_QUERY:
    def __init__(self, XML_FILES):
        self.XML_FILES = XML_FILES
        self.CODE = self.EXTRACT_POWER_QUERY()

    def EXTRACT_POWER_QUERY(self):
        POWER_QUERY_DATA    = []
        CUSTOM_XML          = 'customXml/'
        
        for FILE_NAME, XML_CONTENT in self.XML_FILES.items():
            if FILE_NAME.startswith(CUSTOM_XML) and FILE_NAME.endswith('.xml'):
                try:
                    ROOT        = ET.fromstring(XML_CONTENT)
                    PQ_CODE     = ET.tostring(ROOT, encoding='unicode')
                    POWER_QUERY_DATA.append({   'FILE'      : FILE_NAME,
                                                'PQ_CODE'   : PQ_CODE.strip()})
                
                except ET.ParseError: print(f"Error parsing {FILE_NAME}")
        
        return pd.DataFrame(POWER_QUERY_DATA)


class QUERY_META_DATA:
    def __init__(self, XML_FILES):
        self.XML_FILES      = XML_FILES
        self.META           = self.EXTRACT_META()

    def EXTRACT_META(self):
        CONNECTIONS_XML = 'xl/connections.xml'
        META_DATA = []
        
        if CONNECTIONS_XML in self.XML_FILES:
            ROOT            = ET.fromstring(self.XML_FILES[CONNECTIONS_XML])
            NS              = {'n': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            for CONNECTION in ROOT.findall('.//n:connection', NS):
                NAME        = CONNECTION.get('name')
                CONN_STRING = CONNECTION.find('n:connectionString', NS).text if CONNECTION.find('n:connectionString', NS) else ''
                META_DATA.append({'QUERY_NAME': NAME, 'CONNECTION_STRING': CONN_STRING})
        
        return pd.DataFrame(META_DATA)






class EXCEL_DATA_PARSER:


    def __init__(self, FILE_PATH, LABELS):
        self.XML_EXTRACTOR      = EXCEL_XML_EXTRACTOR(FILE_PATH)
        self.SHEET_NAMES        = WORKBOOK_PARSING(self.XML_EXTRACTOR.XML_FILES, LABELS)
        self.WORKSHEET_PARSER   = WORKSHEET_PARSING(self.XML_EXTRACTOR.XML_FILES, self.SHEET_NAMES.DICT)
        self.CALC_CHAIN         = CALC_CHAIN_PARSING(self.XML_EXTRACTOR.XML_FILES, self.SHEET_NAMES.DICT)
        self.SHEET_RELS         = RELATIVE_SHEETS_PARSING(self.XML_EXTRACTOR.XML_FILES, self.SHEET_NAMES.DICT)
        self.SHARED_STRINGS     = SHARED_STRING_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.STYLES             = STYLES_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.DRAWINGS           = DRAWING_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.TABLES             = TABLES_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.THEME              = THEME_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.CONTENT_TYPE       = CONTENT_TYPE_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.WORKBOOK_RELS      = RELATIVE_WORKBOOK_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.PIVOT              = PIVOT_DATA(FILE_PATH)
        self.MODEL              = EXTRACT_DATA_MODELS(self.XML_EXTRACTOR.XML_FILES)
        self.EXTERNAL_CONN      = EXTERNAL_CONNECTIONS_PARSING(self.XML_EXTRACTOR.XML_FILES)
        self.SHEET_ALT          = SHEET_NAME_MAP(FILE_PATH)
        self.VBA                = VBA_CODE(FILE_PATH)
        self.PWQ                = POWER_QUERY(self.XML_EXTRACTOR.XML_FILES)
        self.META               = QUERY_META_DATA(self.XML_EXTRACTOR.XML_FILES)
        


    def GET_DATA(self):

        return {'WORKSHEETS'        : self.WORKSHEET_PARSER.DICT,
                'SHEET_NAME_MAP'    : self.SHEET_NAMES.DICT,
                'SHARED_STRINGS'    : self.SHARED_STRINGS.DICT,
                'STYLES'            : self.STYLES.DICT,
                'CALC_CHAIN'        : self.CALC_CHAIN.DICT,
                'DRAWINGS'          : self.DRAWINGS.DICT,
                'TABLES'            : self.TABLES.DICT,
                'SHEET_RELS'        : self.SHEET_RELS.DICT,
                'THEME'             : self.THEME.DICT,
                'CONTENT_TYPES'     : self.CONTENT_TYPE.DICT,
                'WORKBOOK_RELS'     : self.WORKBOOK_RELS.DICT, 
                'PIVOTS'            : self.PIVOT.DICT,
                'DATA_MODELS'       : self.MODEL.DICT,
                'EXTERNAL_CONN'     : self.EXTERNAL_CONN.DICT,
                'XML'               : self.XML_EXTRACTOR.XML_FILES,
                'VBA'               : self.VBA.CODE,
                'PWQ'               : self.PWQ.CODE}





class EXTRACT_FORMULAS:
    def __init__(self, WORKSHEETS):
        self.WORKSHEETS = WORKSHEETS
        self.DICT       = self.EXTRACT()

    def EXTRACT(self):

        FORMULAS = {}
        for SHEET_NAME, CELLS in self.WORKSHEETS.items():
            SHEET_FORMULAS = {}
            for CELL_REF, CELL_DATA in CELLS.items():
                if 'FORMULA' in CELL_DATA and CELL_DATA['FORMULA'] is not None:
                    FORMULA = CELL_DATA['FORMULA']
                    if any(err in FORMULA for err in ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A']):    SHEET_FORMULAS[CELL_REF] = {'FORMULA': FORMULA, 'ERROR': True}
                    else:                                                                                   SHEET_FORMULAS[CELL_REF] = {'FORMULA': FORMULA, 'ERROR': False}

            FORMULAS[SHEET_NAME] = SHEET_FORMULAS

        return FORMULAS
    

class CROSS_SHEET_REFERENCE_IDENTIFIER:
    def __init__(self, FORMULAS):
        self.FORMULAS = FORMULAS
        self.DICT = self.IDENTIFY()

    def IDENTIFY(self):
        CROSS_REFERENCES = {}

        # Regex pattern to identify sheet and cell references
        PATTERN = r'(\w+!)?([A-Z]+[0-9]+|[A-Z]+:[A-Z]+)'

        for SHEET_NAME, SHEET_FORMULAS in self.FORMULAS.items():
            # Initialize the sheet's reference list in CROSS_REFERENCES
            CROSS_REFERENCES[SHEET_NAME] = {}

            for CELL_REF, FORMULA_INFO in SHEET_FORMULAS.items():
                FORMULA = FORMULA_INFO['FORMULA']
                MATCHES = re.findall(PATTERN, FORMULA)
                
                if MATCHES:
                    REFERENCES = []
                    for match in MATCHES:
                        if match[0]:  # If the match contains a sheet name
                            REFERENCES.append({'SHEET': match[0].rstrip('!'), 'CELL': match[1]})
                        else:  # Intra-sheet reference; use the current SHEET_NAME
                            REFERENCES.append({'SHEET': SHEET_NAME, 'CELL': match[1]})
                    
                    # Add the references to the specific cell in CROSS_REFERENCES
                    CROSS_REFERENCES[SHEET_NAME][CELL_REF] = REFERENCES

        return CROSS_REFERENCES
    


class ADD_REFERENCES_TO_EXTRACTOR:

    def __init__(self, extractor_dict, cross_ref_dict):
        self.EXTRACTOR_DICT = extractor_dict
        self.CROSS_REF_DICT = cross_ref_dict

    def ADD_REFERENCES(self):
        for SHEET_NAME, CELLS in self.EXTRACTOR_DICT.items():
            for CELL_REF, CELL_DATA in CELLS.items():

                REFERENCES = []
                if SHEET_NAME in self.CROSS_REF_DICT and CELL_REF in self.CROSS_REF_DICT[SHEET_NAME]:
                    REFERENCES = self.CROSS_REF_DICT[SHEET_NAME][CELL_REF]
                
                CELL_DATA['REFERENCES'] = REFERENCES






class FORMULA_ANALYZER:
    def __init__(self, WORKSHEETS, FUNC_PATH):
        self.WORKSHEETS = WORKSHEETS
        self.FUNC_PATH  = FUNC_PATH


    def ANALYSE_FORMULAS(self):
        """
        Loop through each sheet and cell to find and parse formulas.
        Adds a breakdown of each formula to the cell data.
        """

        FUNCTIONS_DF, OPERATORS_DF = self.SOURCE_REFERENCE_INFORMATION(self.FUNC_PATH)

        for SHEET_NAME, CELLS in self.WORKSHEETS.items():
            for CELL_REFS, CELL_DATA in CELLS.items():
                FORMULA = CELL_DATA.get('FORMULA')
                if FORMULA:  
                    FORMULA_BREAKDOWN = self.FORMULA_CONVERSION(('='+FORMULA), OPERATORS_DF, FUNCTIONS_DF)
                    CELL_DATA['FORMULA_BREAKDOWN'] = FORMULA_BREAKDOWN



    def LOAD_TEST_DATA(self, EXCEL_FUNCS_PATH):
        TEST_WB                             = openpyxl.load_workbook(EXCEL_FUNCS_PATH)
        TEST_DICT_WB                        = {'SHEETS': {}, 'DATAFRAME': {}, 'COLUMN_MAP': {}}

        for idx, TEST_SHEET_NAME in enumerate(TEST_WB.sheetnames):
            TEST_WS                         = TEST_WB[TEST_SHEET_NAME]
            TEST_DF                         = pd.DataFrame(list(TEST_WS.values)[1:], columns=list(TEST_WS.values)[0])
            COLUMN_MAPPING_DF               = pd.DataFrame({'LETTER' : [openpyxl.utils.get_column_letter(i + 1) for i in range(TEST_WS.max_column)],
                                                            'HEADER' : list(TEST_WS.values)[0]})

            TEST_DICT_WB['SHEETS'][idx]     = TEST_SHEET_NAME
            TEST_DICT_WB['DATAFRAME'][idx]  = TEST_DF
            TEST_DICT_WB['COLUMN_MAP'][idx] = COLUMN_MAPPING_DF

        return TEST_DICT_WB



    def SOURCE_REFERENCE_INFORMATION(self, EXCEL_FUNCS_PATH):

        FUNC_DATA                       = self.LOAD_TEST_DATA(EXCEL_FUNCS_PATH)
        FUNCTIONS_DF                    = FUNC_DATA['DATAFRAME'][0]
        OPERATORS_DF                    = FUNC_DATA['DATAFRAME'][1]
        FUNCTIONS_DF['JSON_FIELDS']     = FUNCTIONS_DF['JSON_FIELDS'].apply(ast.literal_eval)
        FUNCTIONS_DF['CLUSTERS']        = FUNCTIONS_DF['CLUSTERS'].apply(ast.literal_eval)

        return FUNCTIONS_DF, OPERATORS_DF



    def FORMULA_CONVERSION(self, FORMULA, OPERATORS_DF, FUNCTIONS_DF):

        FORMULA_LIST        = self.EXTRACTION(FORMULA)
        FORMULA_CLEAN       = self.APPLY_OPERATORS(FORMULA, OPERATORS_DF)        
        RESULT              = self.PROCESS_FORMULA(FORMULA_CLEAN[1:], FUNCTIONS_DF, FORMULA_LIST)

        return RESULT



    def EXTRACTION(self, FORMULA):
        FORMULA = FORMULA.strip()

        PATTERN                 = r"([A-Z]+)\("
        FUNCTION_LIST           = re.findall(PATTERN, FORMULA)

        if not FUNCTION_LIST:   return ['HARDCODED']
        return FUNCTION_LIST




    def APPLY_OPERATORS(self, REFERENCE, OPERATORS_DF):

        INITIAL_EQUALS                  = REFERENCE.startswith('=')
        if INITIAL_EQUALS: REFERENCE    = REFERENCE[1:] 


        for idx, ROW in OPERATORS_DF.iterrows():
            EXCEL_OP                    = re.escape(ROW['EXCEL_OPERATOR'])
            PLACEHOLDER                 = ROW['PLACEHOLDER']
            REFERENCE                   = REFERENCE.replace(EXCEL_OP, PLACEHOLDER)
        

        for idx, ROW in OPERATORS_DF.iterrows():
            PLACEHOLDER                 = ROW['PLACEHOLDER']
            PYTHON_OP                   = ROW['PYTHON_OPERATOR']
            REFERENCE                   = REFERENCE.replace(PLACEHOLDER, PYTHON_OP)
        
        if INITIAL_EQUALS: REFERENCE    = '=' + REFERENCE

        return REFERENCE




    def PROCESS_FORMULA(self, FORMULA, FUNCTIONS_DF, FORMULA_LIST):

        TOP_FUNC_NAME, TOP_FUNC_ARGS    = self.PARSE_FUNCTIONS(FORMULA)
        if 'HARDCODED' in FORMULA_LIST: return self.JSON_OUTPUT_HARDCODED(FORMULA.strip('=').strip())
        if TOP_FUNC_NAME not in FUNCTIONS_DF['EXCEL_FUNCTION'].values:
            return {"FUNCTION"      : "UNKNOWN",
                    "COMPONENTS"    : { "FUNCTION_NAME" : TOP_FUNC_NAME,
                                        "ARGUMENTS"     : TOP_FUNC_ARGS}}
        
        if not TOP_FUNC_NAME:           return {}

        return self.JSON_OUTPUT(TOP_FUNC_NAME, TOP_FUNC_ARGS, FUNCTIONS_DF)


    def JSON_OUTPUT_HARDCODED(self, value):
        return {"FUNCTION"      : "HARDCODED",
                "COMPONENTS"    : {"STRING" : value}
                }



    def PARSE_FUNCTIONS(self, FORMULA):
        PATTERN             = r"([A-Z]+)\((.*)\)"
        MATCH               = re.match(PATTERN, FORMULA.strip())
        
        if MATCH:
            FUNCTION_NAME   = MATCH.group(1)
            ARGUMENT_STR    = MATCH.group(2).strip()
            
            ARGUMENTS       = self.SPLIT_ARGUMENTS(ARGUMENT_STR)
            
            return FUNCTION_NAME, ARGUMENTS
        
        return None, []




    def JSON_OUTPUT(self, FUNCTION_NAME, ARGUMENTS, FUNCTIONS_DF):
        JSON_FIELDS                                             = (FUNCTIONS_DF[FUNCTIONS_DF['EXCEL_FUNCTION']==FUNCTION_NAME].reset_index().iloc[:,1:]).at[0, 'JSON_FIELDS']

        if len(ARGUMENTS) == len(JSON_FIELDS):
            COMPONENTS = {}
            
            for i in range(len(ARGUMENTS)):
                NESTED_FUNCS_MATCH                              = re.match(r"([A-Z]+)\((.*)\)", ARGUMENTS[i].strip())
                
                if NESTED_FUNCS_MATCH:
                    NESTED_FUNCS_NAME                           = NESTED_FUNCS_MATCH.group(1)
                    NESTED_FUNCS_ARGS                           = self.SPLIT_ARGUMENTS(NESTED_FUNCS_MATCH.group(2))
                    COMPONENTS[JSON_FIELDS[i]]                  = self.JSON_OUTPUT(NESTED_FUNCS_NAME, NESTED_FUNCS_ARGS, FUNCTIONS_DF)

                else:
                    if JSON_FIELDS[i] == "CONDITION":           COMPONENTS[JSON_FIELDS[i]] = self.PROCESS_CONDITIONS(ARGUMENTS[i].strip())
                    else:                                       COMPONENTS[JSON_FIELDS[i]] = self.NONE_PREFIX(ARGUMENTS[i].strip())
            
            return {"FUNCTION": FUNCTION_NAME, "COMPONENTS": COMPONENTS}
        
        else:    
            return {"FUNCTION": FUNCTION_NAME, "COMPONENTS": ARGUMENTS}



    def SPLIT_ARGUMENTS(self, ARGUMENT_STR):
        ARGS, CURRENT_ARG, PARENTHESIS_COUNT    = [], "", 0

        for CHAR in ARGUMENT_STR:
            if CHAR == ',' and PARENTHESIS_COUNT == 0:
                ARGS.append(CURRENT_ARG.strip())
                CURRENT_ARG = ""

            else:
                CURRENT_ARG += CHAR
                if CHAR == '(':         PARENTHESIS_COUNT += 1
                elif CHAR == ')':       PARENTHESIS_COUNT -= 1

        if CURRENT_ARG:                 ARGS.append(CURRENT_ARG.strip())
        
        return ARGS




    def NONE_PREFIX(self, VALUE):

        REFERENCE_PATTERN   = r"^[A-Za-z]+\d+$"
        RANGE_PATTERN       = r"^[A-Za-z]+:[A-Za-z]+$"
        
        if re.match(REFERENCE_PATTERN, VALUE) or re.match(RANGE_PATTERN, VALUE):

            if "!" not in VALUE:
                return f"NONE!{VALUE}"
            
        return VALUE



    def PROCESS_CONDITIONS(self, CONDITIONS):

        CONDITION_PARTS     = re.split(r'(==|!=|<=|>=|<|>|=)', CONDITIONS)
        PROCESSED_PARTS     = [self.NONE_PREFIX(part.strip()) for part in CONDITION_PARTS]
        
        return ''.join(PROCESSED_PARTS)






def DEPENDENCY_GRAPHING(FORMULAS):

    G = nx.DiGraph()

    for SHEET_NAME, CELLS in FORMULAS.items():
        for CELL, CELL_DATA in CELLS.items():
            NODE_ID = f"{SHEET_NAME}!{CELL}"
            G.add_node(NODE_ID)

            # Add edges for each reference in the formula
            REFERENCES = CELL_DATA.get('REFERENCES', [])
            for REF in REFERENCES:
                REF_SHEET = REF['SHEET']
                REF_CELL = REF['CELL']
                REF_NODE_ID = f"{REF_SHEET}!{REF_CELL}"
                G.add_node(REF_NODE_ID)  
                G.add_edge(REF_NODE_ID, NODE_ID)

    return G




def FORMULA_CLUSTERS(FORMULA):
    
    CLUSTERS = defaultdict(dict)

    for SHEET_NAME, CELLS in FORMULA.items():
        FORM_CLUSTERS = defaultdict(list)

        for CELL, CELL_DATA in CELLS.items():
            FORMULA_KEY = CELL_DATA.get('FORMULA', 'NO_FORMULA')
            FORM_CLUSTERS[FORMULA_KEY].append(CELL)

        CLUSTERS[SHEET_NAME] = dict(FORM_CLUSTERS)

    return CLUSTERS




def MAP_CROSS_REFERENCES(CROSS_SHEET_REF, WORKSHEETS):

    CROSS_SHEET_MAP = defaultdict(list)

    for SHEET_NAME, CELLS in CROSS_SHEET_REF.items():
        for CELL, REFERENCES in CELLS.items():
            for REF in REFERENCES:
                REF_SHEET   = REF['SHEET']
                REF_CELL    = REF['CELL']

                if REF_SHEET in WORKSHEETS:
                    CROSS_SHEET_MAP[(SHEET_NAME, CELL)].append((REF_SHEET, REF_CELL))

    return CROSS_SHEET_MAP







def PHASE_3_OVERVIEW(FORMULA_REFS, CROSS_SHEET_REFS):

    DEPENDENCY_GRAPH    = DEPENDENCY_GRAPHING(FORMULA_REFS)
    FORMULA_CLUSTER     = FORMULA_CLUSTERS(FORMULA_REFS)
    CROSS_SHEET_MAP     = MAP_CROSS_REFERENCES(CROSS_SHEET_REFS, FORMULA_REFS)

    return {'DEPENDENCY_GRAPH'  : DEPENDENCY_GRAPH,
            'CLUSTERS'          : FORMULA_CLUSTER,
            'CROSS_SHEET_MAP'   : CROSS_SHEET_MAP}




def ENHANCED_CLUSTER(FUNC_PATH):
    FUNCTIONS_DF, OPERATORS_DF = FORMULA.SOURCE_REFERENCE_INFORMATION(FUNC_PATH) 
    CLUSTER_INFO = FUNCTIONS_DF.set_index('EXCEL_FUNCTION')['CLUSTERS'].to_dict()
    return CLUSTER_INFO









def CREATE_NESTED_OUTPUT(EXTRACTOR_DICT, FUNC_PATH):
    def STANDARDIZE_CELL_REFERENCE(CELL_REF):
        PATTERN = r'([A-Za-z0-9_]+!)?([A-Za-z]+)([0-9]+)'
        STANDARDIZED_REF = re.sub(PATTERN, r'\1\2*', CELL_REF)
        return STANDARDIZED_REF

    def STANDARDIZE_FORMULA(FORMULA_BREAKDOWN, CLUSTERING_INFO):
        if not FORMULA_BREAKDOWN: return {"FUNCTION": "NO_FUNCTION", "COMPONENTS": {}}

        FUNCTION_NAME           = FORMULA_BREAKDOWN.get('FUNCTION', 'NO_FUNCTION')
        COMPONENTS              = FORMULA_BREAKDOWN.get('COMPONENTS', {})
        CLUSTERING_ATTRIBUTES   = CLUSTERING_INFO.get(FUNCTION_NAME, [])
        STANDARDIZED_COMPONENTS = {}

        for COMP_NAME, COMP_VALUE in COMPONENTS.items():
            if isinstance(COMP_VALUE, dict) and 'FUNCTION' in COMP_VALUE: STANDARDIZED_COMPONENTS[COMP_NAME] = STANDARDIZE_FORMULA(COMP_VALUE, CLUSTERING_INFO)
            else:
                if COMP_NAME in CLUSTERING_ATTRIBUTES: STANDARDIZED_COMPONENTS[COMP_NAME] = STANDARDIZE_CELL_REFERENCE(COMP_VALUE)
                else: STANDARDIZED_COMPONENTS[COMP_NAME] = COMP_VALUE

        return {"FUNCTION": FUNCTION_NAME, "COMPONENTS": STANDARDIZED_COMPONENTS}

    def CONVERT_TO_STANDARDIZED_STRING(NESTED_DICT):
        FUNCTION_NAME           = NESTED_DICT.get('FUNCTION', 'NO_FUNCTION')
        COMPONENTS              = NESTED_DICT.get('COMPONENTS', {})
        # Convert COMPONENTS to a JSON-formatted string
        COMPONENTS_STRING       = json.dumps(COMPONENTS, separators=(',', ':'))
        STANDARDIZED_STRING     = f"{FUNCTION_NAME}|{COMPONENTS_STRING}"
        return STANDARDIZED_STRING

    CLUSTERING_INFO = ENHANCED_CLUSTER(FUNC_PATH)

    NESTED_OUTPUT = {}
    for SHEET_NAME, CELLS in EXTRACTOR_DICT.items():
        NESTED_OUTPUT[SHEET_NAME] = {}
        for CELL_REF, CELL_DATA in CELLS.items():
            FORMULA_BREAKDOWN = CELL_DATA.get('FORMULA_BREAKDOWN')
            if FORMULA_BREAKDOWN:
                NESTED_DICT                         = STANDARDIZE_FORMULA(FORMULA_BREAKDOWN, CLUSTERING_INFO)
                NESTED_STRING                       = CONVERT_TO_STANDARDIZED_STRING(NESTED_DICT)
                NESTED_OUTPUT[SHEET_NAME][CELL_REF] = { 'NESTED_DICT': NESTED_DICT,
                                                        'NESTED_STRING': NESTED_STRING}

    return NESTED_OUTPUT


def GROUP_NESTED_OUTPUT_BY_STRUCTURE(NESTED_OUTPUT):
    CLUSTERS = {}

    for SHEET_NAME, CELLS in NESTED_OUTPUT.items():
        FORM_CLUSTERS = {}

        for CELL, CELL_DATA in CELLS.items():
            NESTED_STRING = CELL_DATA.get('NESTED_STRING', 'NO_NESTED_STRING')
            if NESTED_STRING not in FORM_CLUSTERS:
                FORM_CLUSTERS[NESTED_STRING] = []
            FORM_CLUSTERS[NESTED_STRING].append(CELL)

        CLUSTERS[SHEET_NAME] = FORM_CLUSTERS

    return CLUSTERS


def MAP_SHARED_STRINGS_TO_WORKSHEETS(PARSED_DATA):
    WORKSHEETS = PARSED_DATA['WORKSHEETS']
    SHARED_STRINGS = PARSED_DATA['SHARED_STRINGS']

    for SHEET_NAME, SHEET_DATA in WORKSHEETS.items():
        for CELL_REF, CELL_DATA in SHEET_DATA.items():
            if CELL_DATA.get('TYPE') == 's':  # Shared string type
                SHARED_STRING_INDEX = CELL_DATA.get('VALUE')
                if SHARED_STRING_INDEX is not None:
                    # Map the shared string index to the actual shared string
                    CELL_DATA['VALUE'] = SHARED_STRINGS[int(SHARED_STRING_INDEX)]

    return WORKSHEETS


def TABLES_TO_SHEET(PARSED_DATA, UPDATED_WORKSHEETS):

    for TABLE_FILE, TABLE_DATA in PARSED_DATA['TABLES'] .items():
        LOCATION    = TABLE_DATA.get('SHEET_LOCATION') 
        COLS        = TABLE_DATA.get('COLUMNS', [])  

        if LOCATION:
            COL_START, ROW_START = re.split(r'(\d+)', LOCATION)[:2]
            ROW_START = int(ROW_START)

            for SHEET_NAME, SHEET_DATA in UPDATED_WORKSHEETS.items():
                MATCHING_COLS = []

                for CELL_REF, CELL_DATA in SHEET_DATA.items():
                    COL_REF, ROW_REF = re.split(r'(\d+)', CELL_REF)[:2]
                    ROW_REF = int(ROW_REF)

                    if ROW_REF == ROW_START:
                        CELL_VALUE = CELL_DATA.get('VALUE', None)
                        if CELL_VALUE in COLS:
                            MATCHING_COLS.append(CELL_VALUE)

                if len(MATCHING_COLS) == len(COLS):
                    TABLE_DATA['LOCATION'] = SHEET_NAME 
                    break

    return PARSED_DATA 




def IDENTIFY_TABLES_WITH_HEADERS_AND_GAPS(WORKSHEETS, MAX_COLUMN_GAP=2):
    TABLES = {}

    for SHEET_NAME, CELLS in WORKSHEETS.items():
        SHEET_TABLES = []
        CURRENT_TABLE = None
        PREVIOUS_COL_INDEX = None

        for CELL_REF, CELL_DATA in sorted(CELLS.items()):
            CELL_VALUE = CELL_DATA.get('VALUE')
            FORMULA = CELL_DATA.get('FORMULA')

            if CELL_VALUE or FORMULA:
                COL, ROW = re.split(r'(\d+)', CELL_REF)[:2]
                ROW = int(ROW)
                COL_INDEX = ord(COL.upper()) - ord('A')

                if PREVIOUS_COL_INDEX is not None and (COL_INDEX - PREVIOUS_COL_INDEX) > MAX_COLUMN_GAP:
                    if CURRENT_TABLE and CURRENT_TABLE['CELLS']:
                        SHEET_TABLES.append(CURRENT_TABLE)
                    CURRENT_TABLE = None

                if CURRENT_TABLE is None:
                    CURRENT_TABLE = {
                        'START_CELL': CELL_REF,
                        'HEADER': [],
                        'ROWS': 0,
                        'COLUMNS': 0,
                        'CELLS': []
                    }

                CURRENT_TABLE['CELLS'].append(CELL_REF)
                PREVIOUS_COL_INDEX = COL_INDEX

        if CURRENT_TABLE and CURRENT_TABLE['CELLS']:
            SHEET_TABLES.append(CURRENT_TABLE)

        for TABLE in SHEET_TABLES:
            FIRST_ROW_CELLS = [CELL_REF for CELL_REF in TABLE['CELLS'] if re.search(r'1$', CELL_REF)]
            for CELL_REF in FIRST_ROW_CELLS:
                CELL_DATA = CELLS.get(CELL_REF, {})
                CELL_VALUE = CELL_DATA.get('VALUE')
                FORMULA = CELL_DATA.get('FORMULA')

                if CELL_VALUE and not FORMULA:
                    TABLE['HEADER'].append(CELL_VALUE)

            ROWS = set()
            COLS = set()
            for CELL in TABLE['CELLS']:
                COL, ROW = re.split(r'(\d+)', CELL)[:2]
                ROW = int(ROW)
                COL_INDEX = ord(COL.upper()) - ord('A')
                ROWS.add(ROW)
                COLS.add(COL_INDEX)
            TABLE['ROWS'] = len(ROWS)
            TABLE['COLUMNS'] = len(COLS)

        if SHEET_TABLES:
            TABLES[SHEET_NAME] = SHEET_TABLES

    return TABLES







def FULL_RUN(FILE_PATH, FUNC_PATH, LABELS):


    EXCEL_PARSED            = EXCEL_DATA_PARSER(FILE_PATH, LABELS)
    PARSED_DATA             = EXCEL_PARSED.GET_DATA()

    EXTRACTOR               = EXTRACT_FORMULAS(PARSED_DATA['WORKSHEETS'])
    FORMULA                 = FORMULA_ANALYZER(EXTRACTOR.DICT, FUNC_PATH)

    FORMULA.ANALYSE_FORMULAS() 

    CROSS_REF               = CROSS_SHEET_REFERENCE_IDENTIFIER(EXTRACTOR.DICT)
    REFERENCES              = ADD_REFERENCES_TO_EXTRACTOR(EXTRACTOR.DICT, CROSS_REF.DICT)

    REFERENCES.ADD_REFERENCES()

    P3_RESULTS              = PHASE_3_OVERVIEW(EXTRACTOR.DICT, CROSS_REF.DICT)
    NESTED_OUTPUT           = CREATE_NESTED_OUTPUT(EXTRACTOR.DICT, FUNC_PATH)
    CLUSTERED_NESTED_OUTPUT = GROUP_NESTED_OUTPUT_BY_STRUCTURE(NESTED_OUTPUT)

    UPDATED_WORKSHEETS      = MAP_SHARED_STRINGS_TO_WORKSHEETS(PARSED_DATA)
    PARSED_DATA             = TABLES_TO_SHEET(PARSED_DATA, UPDATED_WORKSHEETS)
    OUTPUT_DETECTED_TABLES  = IDENTIFY_TABLES_WITH_HEADERS_AND_GAPS(UPDATED_WORKSHEETS)

    return PARSED_DATA, EXTRACTOR.DICT, NESTED_OUTPUT, CLUSTERED_NESTED_OUTPUT, UPDATED_WORKSHEETS, OUTPUT_DETECTED_TABLES
 